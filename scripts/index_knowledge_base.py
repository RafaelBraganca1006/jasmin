"""
Indexa knowledge_base/ no ChromaDB. Idempotente: recria as collections a cada
execução. Rode UMA vez (e de novo se os documentos mudarem):

    python scripts/index_knowledge_base.py

Duas collections:
  tuss_procedures  — uma linha por procedimento (xlsx + PDFs-tabela), com
                     codigo_tuss em metadata para lookup exato.
  dental_rules     — blocos de texto de regras/legislação (PDFs).

Estratégia (ver discussão de arquitetura):
  - Tabelas são chunked POR LINHA, com os cabeçalhos repetidos em cada chunk
    (embeddings densos não entendem linha solta sem rótulo).
  - codigo_tuss vai para metadata → o servidor faz lookup EXATO por código via
    filtro de metadata (vector search erra em código alfanumérico).
"""
import os
import re
import sys
import time

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import rag_common as cfg  # noqa: E402

# Windows: console é cp1252 e quebra em "→"/"─". Força UTF-8 no stdout.
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

# ── Classificação dos documentos ────────────────────────────────────────────
# PDFs cujo conteúdo é majoritariamente TABELA de procedimentos → tuss_procedures
TUSS_TABLE_PDFS = {
    "tabela_tuss_odontologica.pdf",
    "TABELA-_Odontologica_Fiosaude_2024_MAR2024-FEV2025.pdf",
}
# PDFs de regra/legislação (texto corrido) → dental_rules
RULES_PDFS = {
    "RN-ANS-de-2021-no-465_Atualiza-Rol-de-Procedimentos.pdf",
    "Tabela_Procedimentos_Odontologicos_Edital_Chamamento_26.pdf",
}

# Código TUSS: 8 dígitos (ex.: 81000030) ou pontilhado (ex.: 8.10.00.03.0).
RE_TUSS = re.compile(r"\b(\d{8}|\d(?:\.\d{2}){3}\.\d)\b")
# Sinais textuais de "exige X" nas tabelas/regras.
RE_SIM = re.compile(r"\b(sim|obrigat)", re.IGNORECASE)


def norm(s):
    return ("" if s is None else str(s)).strip()


def find_tuss_code(text):
    m = RE_TUSS.search(text or "")
    return m.group(1).replace(".", "") if m else ""


# ── XLSX ─────────────────────────────────────────────────────────────────────
def detect_header_row(ws, scan=12):
    """Acha a linha (0-based) cujo conteúdo parece um cabeçalho (tem 'código')."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=scan, values_only=True)):
        cells = [norm(c).lower() for c in row]
        if any(c.startswith("c") and "digo" in c for c in cells):  # "código"/"codigo"
            return i, [norm(c) for c in row]
    return 0, [norm(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]


def index_xlsx(path, fname, docs):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]  # 1ª planilha é a tabela principal nos dois arquivos
    header_idx, headers = detect_header_row(ws)
    headers = [h if h else f"col{j}" for j, h in enumerate(headers)]

    # Coluna de código e coluna "OD" (odonto) na tabela de correlação ROL.
    code_col = next((j for j, h in enumerate(headers) if h.lower().startswith("c") and "digo" in h.lower()), 0)
    od_col = next((j for j, h in enumerate(headers) if h.strip().upper() == "OD"), None)

    count = 0
    for r, row in enumerate(ws.iter_rows(min_row=header_idx + 2, values_only=True)):
        cells = [norm(c) for c in row]
        if not any(cells):
            continue
        code = norm(cells[code_col]) if code_col < len(cells) else ""
        code = find_tuss_code(code) or find_tuss_code(" ".join(cells))
        if not code:
            continue
        # Na correlação ROL (6700+ linhas, maioria médica), só odontologia (OD preenchido).
        if od_col is not None and od_col < len(cells) and not cells[od_col]:
            continue
        parts = [f"{headers[j]}: {cells[j]}" for j in range(min(len(headers), len(cells))) if cells[j]]
        content = " | ".join(parts)
        docs.append({
            "id": f"{fname}#r{r}",
            "content": content,
            "metadata": {"source": fname, "row_index": r, "type": "xlsx", "codigo_tuss": code},
        })
        count += 1
    wb.close()
    return count


# ── PDF tabela → linhas ──────────────────────────────────────────────────────
def index_pdf_table(path, fname, docs):
    import fitz
    doc = fitz.open(path)
    count = 0
    for pno in range(doc.page_count):
        page = doc[pno]
        rows = []
        try:
            tables = page.find_tables()
            for tab in tables.tables:
                ext = tab.extract()
                if ext:
                    rows.extend(ext)
        except Exception:
            rows = []

        if rows:
            for ri, cells in enumerate(rows):
                cells = [norm(c) for c in cells]
                joined = " | ".join(c for c in cells if c)
                code = find_tuss_code(joined)
                if not code or len(joined) < 8:
                    continue
                meta = {"source": fname, "page": pno + 1, "type": "pdf-table", "codigo_tuss": code}
                low = joined.lower()
                if "rx" in low or "radiograf" in low:
                    meta["menciona_rx"] = True
                if "autoriz" in low:
                    meta["menciona_autorizacao"] = True
                docs.append({
                    "id": f"{fname}#p{pno+1}r{ri}",
                    "content": joined,
                    "metadata": meta,
                })
                count += 1
        else:
            # Página sem tabela detectada: não perder os códigos → chunk por bloco.
            for ci, chunk in enumerate(chunk_blocks(page)):
                code = find_tuss_code(chunk)
                docs.append({
                    "id": f"{fname}#p{pno+1}b{ci}",
                    "content": chunk,
                    "metadata": {"source": fname, "page": pno + 1, "type": "pdf-table",
                                 "codigo_tuss": code} if code else
                                {"source": fname, "page": pno + 1, "type": "pdf-table"},
                })
                count += 1
    doc.close()
    return count


# ── PDF texto → blocos ───────────────────────────────────────────────────────
MAX_CHARS = 2000  # ~500 tokens


def chunk_blocks(page):
    """Junta blocos de texto da página em chunks de ~500 tokens."""
    blocks = [norm(b[4]) for b in page.get_text("blocks") if norm(b[4])]
    chunks, buf = [], ""
    for b in blocks:
        if len(buf) + len(b) + 2 > MAX_CHARS and buf:
            chunks.append(buf.strip())
            buf = ""
        buf += b + "\n"
    if buf.strip():
        chunks.append(buf.strip())
    return chunks


def index_pdf_rules(path, fname, docs):
    import fitz
    doc = fitz.open(path)
    count = 0
    for pno in range(doc.page_count):
        for ci, chunk in enumerate(chunk_blocks(doc[pno])):
            if len(chunk) < 30:
                continue
            docs.append({
                "id": f"{fname}#p{pno+1}c{ci}",
                "content": chunk,
                "metadata": {"source": fname, "page": pno + 1, "type": "pdf", "chunk_index": ci},
            })
            count += 1
    doc.close()
    return count


# ── Escrita no ChromaDB em lotes ─────────────────────────────────────────────
def add_batched(collection, docs, batch=128):
    for i in range(0, len(docs), batch):
        part = docs[i:i + batch]
        collection.add(
            ids=[d["id"] for d in part],
            documents=[d["content"] for d in part],
            metadatas=[d["metadata"] for d in part],
        )


def reset_collection(client, name):
    try:
        client.delete_collection(name)
    except Exception:
        pass
    return client.create_collection(
        name=name,
        embedding_function=cfg.get_embedding_function(),
        metadata={"hnsw:space": cfg.CHROMA_SPACE},
    )


def main():
    t0 = time.time()
    if not os.path.isdir(cfg.KB_DIR):
        print(f"ERRO: {cfg.KB_DIR} não existe.")
        sys.exit(1)

    files = sorted(os.listdir(cfg.KB_DIR))
    print(f"knowledge_base/: {len(files)} arquivos")
    print(f"Modelo de embedding: {cfg.EMBED_MODEL}")
    print("Carregando modelo e abrindo ChromaDB (primeira vez baixa o modelo)...\n")

    client = cfg.get_client()
    col_tuss = reset_collection(client, cfg.COLLECTION_TUSS)
    col_rules = reset_collection(client, cfg.COLLECTION_RULES)

    tuss_docs, rules_docs = [], []
    for fname in files:
        path = os.path.join(cfg.KB_DIR, fname)
        if not os.path.isfile(path):
            continue
        low = fname.lower()
        try:
            if low.endswith(".xlsx"):
                n = index_xlsx(path, fname, tuss_docs)
                print(f"  [xlsx]  {fname}: {n} linhas → tuss_procedures")
            elif fname in TUSS_TABLE_PDFS:
                n = index_pdf_table(path, fname, tuss_docs)
                print(f"  [pdf-t] {fname}: {n} linhas/blocos → tuss_procedures")
            elif low.endswith(".pdf"):
                n = index_pdf_rules(path, fname, rules_docs)
                print(f"  [pdf]   {fname}: {n} chunks → dental_rules")
            else:
                print(f"  [skip]  {fname}")
        except Exception as e:
            print(f"  [ERRO]  {fname}: {e}")

    print("\nGravando no ChromaDB (gera embeddings, pode levar alguns minutos)...")
    add_batched(col_tuss, tuss_docs)
    add_batched(col_rules, rules_docs)

    print("\n── Resultado ──────────────────────────────────────────────")
    print(f"tuss_procedures : {col_tuss.count()} documentos")
    print(f"dental_rules    : {col_rules.count()} documentos")

    # Queries de teste.
    print("\n── Query de teste ─────────────────────────────────────────")
    try:
        q = col_tuss.query(query_texts=["tratamento de canal molar"], n_results=2)
        for d, m, dist in zip(q["documents"][0], q["metadatas"][0], q["distances"][0]):
            print(f"  [tuss] sim={1-dist:.2f} {m.get('source')}: {d[:80]}")
    except Exception as e:
        print(f"  [tuss] erro: {e}")
    try:
        q = col_rules.query(query_texts=["autorização prévia procedimento odontológico"], n_results=2)
        for d, m, dist in zip(q["documents"][0], q["metadatas"][0], q["distances"][0]):
            print(f"  [rules] sim={1-dist:.2f} {m.get('source')} p{m.get('page')}: {d[:80]}")
    except Exception as e:
        print(f"  [rules] erro: {e}")

    print(f"\nConcluído em {time.time()-t0:.1f}s")


if __name__ == "__main__":
    main()
